"""
Document conversion endpoints:
  POST /convert/word-to-pdf
  POST /convert/pdf-to-word
  POST /convert/excel-to-pdf
"""

import uuid
import logging
from pathlib import Path

from fastapi import APIRouter, UploadFile, File, BackgroundTasks, Request, HTTPException

from utils.validators import validate_file
from utils.cleanup import schedule_file_cleanup

router = APIRouter()
logger = logging.getLogger(__name__)

WORD_EXTS = {".docx", ".doc"}
PDF_EXTS  = {".pdf"}
EXCEL_EXTS = {".xlsx", ".xls"}


# ── Helper ────────────────────────────────────────────────────────────────────
def _save_upload(content: bytes, suffix: str, tmp_dir: Path) -> Path:
    input_id = uuid.uuid4().hex
    path = tmp_dir / f"{input_id}{suffix}"
    path.write_bytes(content)
    return path


# ── Word → PDF ────────────────────────────────────────────────────────────────
@router.post("/word-to-pdf", summary="Convert DOCX/DOC → PDF")
async def word_to_pdf(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
):
    """Convert a Word document (.docx) to PDF using python-docx + ReportLab."""
    await validate_file(file, allowed_exts=WORD_EXTS)

    content = await file.read()
    tmp_dir: Path = request.app.state.tmp_dir

    # Save input
    ext = Path(file.filename).suffix.lower()
    input_path = _save_upload(content, ext, tmp_dir)

    # Output file
    output_id = uuid.uuid4().hex
    output_path = tmp_dir / f"{output_id}.pdf"

    try:
        _convert_docx_to_pdf(input_path, output_path)
    except Exception as e:
        input_path.unlink(missing_ok=True)
        logger.error("word-to-pdf failed: %s", e)
        raise HTTPException(status_code=422, detail=f"Conversion failed: {str(e)}")

    # Schedule cleanup for both files
    schedule_file_cleanup(background_tasks, input_path, output_path)

    return {
        "task_id": output_id,
        "file_id": output_id,
        "status": "done",
        "output_filename": f"{Path(file.filename).stem}.pdf",
        "download_url": f"/download/{output_id}",
    }


def _convert_docx_to_pdf(input_path: Path, output_path: Path):
    """
    Convert DOCX → PDF.

    Strategy:
      1. Extract text + basic structure with python-docx.
      2. Render to PDF with ReportLab (cross-platform, no LibreOffice needed).

    For pixel-perfect rendering, swap this for LibreOffice headless:
      subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", str(input_path)])
    """
    from docx import Document
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
    from reportlab.lib import colors

    doc = Document(str(input_path))
    pdf_doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2.5 * cm,
    )

    styles = getSampleStyleSheet()
    story = []

    heading_sizes = {1: 20, 2: 16, 3: 14, 4: 12}

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            story.append(Spacer(1, 0.3 * cm))
            continue

        style_name = para.style.name or ""

        if style_name.startswith("Heading"):
            level = int(style_name[-1]) if style_name[-1].isdigit() else 1
            size = heading_sizes.get(level, 14)
            style = ParagraphStyle(
                f"H{level}",
                fontName="Helvetica-Bold",
                fontSize=size,
                spaceAfter=6,
                spaceBefore=12,
                textColor=colors.HexColor("#1a1a2e"),
            )
            story.append(Paragraph(text, style))
        else:
            style = ParagraphStyle(
                "Body",
                fontName="Helvetica",
                fontSize=10.5,
                leading=15,
                spaceAfter=4,
                alignment=TA_JUSTIFY,
            )
            story.append(Paragraph(text, style))

    pdf_doc.build(story)


# ── PDF → Word ────────────────────────────────────────────────────────────────
@router.post("/pdf-to-word", summary="Convert PDF → DOCX")
async def pdf_to_word(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
):
    """Convert a PDF to a Word document (.docx) using pdf2docx."""
    await validate_file(file, allowed_exts=PDF_EXTS)

    content = await file.read()
    tmp_dir: Path = request.app.state.tmp_dir

    input_path = _save_upload(content, ".pdf", tmp_dir)
    output_id = uuid.uuid4().hex
    output_path = tmp_dir / f"{output_id}.docx"

    try:
        _convert_pdf_to_docx(input_path, output_path)
    except Exception as e:
        input_path.unlink(missing_ok=True)
        logger.error("pdf-to-word failed: %s", e)
        raise HTTPException(status_code=422, detail=f"Conversion failed: {str(e)}")

    schedule_file_cleanup(background_tasks, input_path, output_path)

    return {
        "task_id": output_id,
        "file_id": output_id,
        "status": "done",
        "output_filename": f"{Path(file.filename).stem}.docx",
        "download_url": f"/download/{output_id}",
    }


def _convert_pdf_to_docx(input_path: Path, output_path: Path):
    """Convert PDF to DOCX using pdf2docx library."""
    from pdf2docx import Converter

    cv = Converter(str(input_path))
    try:
        cv.convert(str(output_path), start=0, end=None)
    finally:
        cv.close()


# ── Excel → PDF ───────────────────────────────────────────────────────────────
@router.post("/excel-to-pdf", summary="Convert XLSX/XLS → PDF")
async def excel_to_pdf(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
):
    """Convert an Excel spreadsheet to PDF using openpyxl + ReportLab."""
    await validate_file(file, allowed_exts=EXCEL_EXTS)

    content = await file.read()
    tmp_dir: Path = request.app.state.tmp_dir

    ext = Path(file.filename).suffix.lower()
    input_path = _save_upload(content, ext, tmp_dir)
    output_id = uuid.uuid4().hex
    output_path = tmp_dir / f"{output_id}.pdf"

    try:
        _convert_xlsx_to_pdf(input_path, output_path)
    except Exception as e:
        input_path.unlink(missing_ok=True)
        logger.error("excel-to-pdf failed: %s", e)
        raise HTTPException(status_code=422, detail=f"Conversion failed: {str(e)}")

    schedule_file_cleanup(background_tasks, input_path, output_path)

    return {
        "task_id": output_id,
        "file_id": output_id,
        "status": "done",
        "output_filename": f"{Path(file.filename).stem}.pdf",
        "download_url": f"/download/{output_id}",
    }


def _convert_xlsx_to_pdf(input_path: Path, output_path: Path):
    """Render Excel data into a PDF table with ReportLab."""
    import openpyxl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    wb = openpyxl.load_workbook(str(input_path), data_only=True)
    pdf_doc = SimpleDocTemplate(str(output_path), pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        story.append(Paragraph(f"Sheet: {sheet_name}", styles["Heading2"]))
        story.append(Spacer(1, 0.3 * cm))

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Convert all cells to string
        table_data = [[str(cell) if cell is not None else "" for cell in row] for row in rows]

        col_count = len(table_data[0]) if table_data else 1
        col_width = (landscape(A4)[0] - 4 * cm) / max(col_count, 1)

        tbl = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 0.8 * cm))

    pdf_doc.build(story)
